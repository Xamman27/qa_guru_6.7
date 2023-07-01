from openpyxl import load_workbook
import os


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_for_xlsx():
    path_to_resorces = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'resources')
    workbook = load_workbook(os.path.join(path_to_resorces, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'
